import openpyxl
import json

def map_skills_to_dictionary(filename):
  """
  Maps skills from two subsheets of an xlsx file to a dictionary.

  Args:
      filename: Path to the xlsx file containing subsheets.

  Returns:
      A dictionary with "text" and "labels" keys.
  """
  wb = openpyxl.load_workbook(filename)
  success_profiles = wb["Success Profiles"]
  skills = wb["Skills"]

  # Define empty dictionary
  data = {"text": [], "labels": []}

  # Get unique skills
  unique_skills = set()
  for row in skills.iter_rows(min_row=2):
    job_code = row[0].value
    skill_name = row[5].value
    unique_skills.add(skill_name)

  # Map job descriptions and skills
  rows_index = 0
  for row in success_profiles.iter_rows(min_row=3):
    rows_index += 1
    print("Row", rows_index)
    description = row[5].value
    job_code = row[0].value

    # Split description into sentences
    sentences = description.split(". ")

    # Initialize label vector for each sentence
    labels = [0] * len(unique_skills)

    # Find relevant skills for this job code
    for skill_row in skills.iter_rows(min_row=3):
      skill_job_code = skill_row[0].value
      skill_name = skill_row[5].value
      if skill_job_code == job_code:
        index = list(unique_skills).index(skill_name)
        labels[index] = 1

    # Add sentences and labels to dictionary
    data["text"].extend(sentences)
    data["labels"].extend([labels] * len(sentences))

  return data

def save_to_json(data, filename):
  with open(filename, 'w') as outfile:
    json.dump(data, outfile, indent=4)  # Indent for readability

# Example usage
data = map_skills_to_dictionary("dataset.xlsx")
save_to_json(data, "skills_data.json")  # Adjust filename as needed
